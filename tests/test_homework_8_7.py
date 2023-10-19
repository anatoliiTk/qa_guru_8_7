import zipfile
from PyPDF2 import PdfReader
import xlrd
from openpyxl import load_workbook

from tests.conftest import ZIP_TMP_PATH, PDF_RESOURCES_PATH, XLS_RESOURCES_PATH, XLSX_RESOURCES_PATH, TXT_RESOURCES_PATH


def test_file_pdf():
    with zipfile.ZipFile(ZIP_TMP_PATH, 'r') as testzip:
        with testzip.open('testpdf.pdf') as pdf_t:
            pdf_t = PdfReader(pdf_t)
            page = pdf_t.pages[1]
            text = page.extract_text()
            assert 'Hallo World!!!' in text

            x = testzip.read('testpdf.pdf')
            pdf_file_size = len(x)
            exected_file_size = 184852
            assert pdf_file_size == exected_file_size


def test_xls():
    with zipfile.ZipFile(ZIP_TMP_PATH, 'r') as testzip:
        with testzip.open('testxls.xls') as xls_t:
            xls_content = xls_t.read()
            xls_workbook = xlrd.open_workbook(file_contents=xls_content)
            sheet = xls_workbook.sheet_by_index(0)
            expected_value = 'Hallo World!'
            actual_value = sheet.cell_value(0, 0)
            assert actual_value == expected_value


def test_xlsx():
    with zipfile.ZipFile(ZIP_TMP_PATH, 'r') as myzip:
        with myzip.open('testxlsx.xlsx') as xlsx_t:
            xlsx_file = load_workbook(xlsx_t)
            sheet = xlsx_file.active
            value = sheet.cell(row=3, column=1).value
            assert value == 'Hallo World!!!'

def test_txt():
    with zipfile.ZipFile(ZIP_TMP_PATH, 'r') as myzip:
        with myzip.open('testtxt.txt') as txt_t:
            assert txt_t.read().decode('utf-8') == 'Hello World!\r\nHello World!!'
